from openpyxl import Workbook

source = open("../res/sentences.txt", "r")
result = open("../res/segmented.txt", "r")

good_sentences = source.read()
punkt_sentences = result.read()

source.close()
result.close()

wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1, value='nomor')
ws.cell(row=1, column=2, value='kalimat asli')
ws.cell(row=1, column=3, value='hasil punkt')
err_count = 0
stc_start = 0
stc_end = 0
for char in good_sentences:
	stc_end += 1
	if char == '\n':
		good_stc = good_sentences[stc_start:stc_end]
		punkt_stc = punkt_sentences[stc_start:stc_end]
		if good_stc != punkt_stc:
			err_count += 1
			ws.cell(row=err_count+1, column=1, value=err_count)
			ws.cell(row=err_count+1, column=2, value=good_stc.rstrip())
			ws.cell(row=err_count+1, column=3, value=punkt_stc.rstrip())
		stc_start = stc_end

wb.save('../res/punkt_error.xlsx')